from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


BASE = Path(r"C:\.codex\Playground\output")


def load_looker_json(path: Path) -> dict:
    text = path.read_text(encoding="utf-8").lstrip()
    if text.startswith(")]}'"):
        text = text.split("\n", 1)[1]
    return json.loads(text)


def pct(part: float, total: float) -> float:
    return 0.0 if not total else part / total


def change_pct(current: float, previous: float) -> float:
    return 0.0 if not previous else (current - previous) / previous


def round2(value: float) -> float:
    return round(float(value), 2)


def parse_weekly_rows() -> list[dict]:
    payload = load_looker_json(BASE / "req78_response.txt")
    cols = payload["dataResponse"][0]["dataSubset"][0]["dataset"]["tableDataset"]["column"]
    rows = []
    for idx, week_start in enumerate(cols[0]["dateColumn"]["values"]):
        rows.append(
            {
                "week_start": week_start,
                "impressions": int(cols[1]["longColumn"]["values"][idx]),
                "clicks": int(cols[2]["longColumn"]["values"][idx]),
                "ctr": float(cols[3]["doubleColumn"]["values"][idx]),
                "avg_cpc": float(cols[4]["doubleColumn"]["values"][idx]),
                "cost": float(cols[5]["doubleColumn"]["values"][idx]),
                "conversions": float(cols[6]["doubleColumn"]["values"][idx]),
                "cpa": float(cols[7]["doubleColumn"]["values"][idx]),
            }
        )
    filtered = []
    for row in rows:
        d = date.fromisoformat(row["week_start"])
        if date(2025, 9, 29) <= d <= date(2025, 12, 29):
            row["period_label"] = (
                "opening_overlap_week"
                if row["week_start"] == "2025-09-29"
                else "closing_overlap_week"
                if row["week_start"] == "2025-12-29"
                else "in_quarter_week"
            )
            filtered.append(row)
    filtered.sort(key=lambda r: r["week_start"])
    return filtered


def parse_monthly_rows() -> list[dict]:
    payload = load_looker_json(BASE / "req79_response.txt")
    cols = payload["dataResponse"][2]["dataSubset"][0]["dataset"]["tableDataset"]["column"]
    rows = []
    for idx, month_start in enumerate(cols[0]["dateColumn"]["values"]):
        d = date.fromisoformat(month_start)
        if date(2025, 10, 1) <= d <= date(2025, 12, 1):
            cost = float(cols[1]["doubleColumn"]["values"][idx])
            conversions = float(cols[2]["doubleColumn"]["values"][idx])
            rows.append(
                {
                    "month_start": month_start,
                    "cost": cost,
                    "conversions": conversions,
                    "cpa": 0.0 if not conversions else cost / conversions,
                }
            )
    return rows


def parse_single_table(path: Path, dim_keys: list[str], metric_keys: list[str]) -> tuple[list[dict], dict]:
    payload = load_looker_json(path)
    subset = payload["dataResponse"][0]["dataSubset"]
    main_cols = subset[0]["dataset"]["tableDataset"]["column"]
    totals_cols = subset[1]["dataset"]["tableDataset"]["column"]
    rows = []
    row_count = subset[0]["dataset"]["tableDataset"]["size"]
    all_keys = dim_keys + metric_keys
    for idx in range(row_count):
        row = {}
        for col_idx, key in enumerate(all_keys):
            col = main_cols[col_idx]
            if "stringColumn" in col:
                row[key] = col["stringColumn"]["values"][idx]
            elif "longColumn" in col:
                row[key] = int(col["longColumn"]["values"][idx])
            elif "doubleColumn" in col:
                row[key] = float(col["doubleColumn"]["values"][idx])
            else:
                row[key] = col["dateColumn"]["values"][idx]
        rows.append(row)
    totals = {}
    for idx, key in enumerate(metric_keys):
        col = totals_cols[idx]
        if "longColumn" in col:
            totals[key] = int(col["longColumn"]["values"][0])
        elif "doubleColumn" in col:
            totals[key] = float(col["doubleColumn"]["values"][0])
    return rows, totals


weekly_reports = [
    {
        "week_label": "29.09-05.10",
        "week_start": "2025-09-29",
        "week_end": "2025-10-05",
        "period_label": "opening_overlap_week",
        "actions": "Робота з бюджетом, контроль результатів, регулярний моніторинг.",
        "results": "32 конверсії за тиждень, CPA 35; за місяць 62 конверсії, CPA 69. Один слабший тиждень не зіпсував загальну ефективність.",
        "plan_to_do": "Продовжити оптимізацію та моніторинг.",
        "requests": "",
        "weekly_conversions_reported": 32.0,
        "weekly_cpa_reported": 35.0,
    },
    {
        "week_label": "06.10-12.10",
        "week_start": "2025-10-06",
        "week_end": "2025-10-12",
        "period_label": "in_quarter_week",
        "actions": "Робота з бюджетом, стратегією ставок і ключовими словами.",
        "results": "26 конверсій, CPA 32.",
        "plan_to_do": "Продовжити оптимізацію кампаній.",
        "requests": "",
        "weekly_conversions_reported": 26.0,
        "weekly_cpa_reported": 32.0,
    },
    {
        "week_label": "13.10-19.10",
        "week_start": "2025-10-13",
        "week_end": "2025-10-19",
        "period_label": "in_quarter_week",
        "actions": "Перевірка трекінгу та бюджетів.",
        "results": "7 конверсій, CPA 136. Попит був відчутно слабший, ніж попереднього тижня.",
        "plan_to_do": "Стабілізувати результат і моніторити попит.",
        "requests": "",
        "weekly_conversions_reported": 7.0,
        "weekly_cpa_reported": 136.0,
    },
    {
        "week_label": "20.10-26.10",
        "week_start": "2025-10-20",
        "week_end": "2025-10-26",
        "period_label": "in_quarter_week",
        "actions": "Робота з трекінгом.",
        "results": "CPA 71.",
        "plan_to_do": "Продовжити оптимізацію.",
        "requests": "",
        "weekly_conversions_reported": "",
        "weekly_cpa_reported": 71.0,
    },
    {
        "week_label": "27.10-02.11",
        "week_start": "2025-10-27",
        "week_end": "2025-11-02",
        "period_label": "in_quarter_week",
        "actions": "Оновлення bidding strategy.",
        "results": "29 конверсій, CPA 35; за місяць 92 конверсії, CPA 46. Ознак click fraud не виявлено.",
        "plan_to_do": "Продовжити роботу зі ставками.",
        "requests": "",
        "weekly_conversions_reported": 29.0,
        "weekly_cpa_reported": 35.0,
    },
    {
        "week_label": "03.11-09.11",
        "week_start": "2025-11-03",
        "week_end": "2025-11-09",
        "period_label": "in_quarter_week",
        "actions": "Робота з трекінгом.",
        "results": "13 конверсій, CPA 84.",
        "plan_to_do": "Продовжити оптимізацію.",
        "requests": "",
        "weekly_conversions_reported": 13.0,
        "weekly_cpa_reported": 84.0,
    },
    {
        "week_label": "10.11-16.11",
        "week_start": "2025-11-10",
        "week_end": "2025-11-16",
        "period_label": "in_quarter_week",
        "actions": "Geo exclusions, розширення language targeting, оновлення PMax asset groups.",
        "results": "26 конверсій, CPA 51.",
        "plan_to_do": "Далі чистити трафік і доопрацьовувати PMax.",
        "requests": "",
        "weekly_conversions_reported": 26.0,
        "weekly_cpa_reported": 51.0,
    },
    {
        "week_label": "17.11-23.11",
        "week_start": "2025-11-17",
        "week_end": "2025-11-23",
        "period_label": "in_quarter_week",
        "actions": "Опрацювання negative keywords.",
        "results": "23 конверсії, CPA 48.",
        "plan_to_do": "Продовжити чистку пошукових запитів.",
        "requests": "",
        "weekly_conversions_reported": 23.0,
        "weekly_cpa_reported": 48.0,
    },
    {
        "week_label": "24.11-30.11",
        "week_start": "2025-11-24",
        "week_end": "2025-11-30",
        "period_label": "in_quarter_week",
        "actions": "Робота з трекінгом.",
        "results": "17 конверсій, CPA 64; за місяць 88 конверсій, CPA 56.",
        "plan_to_do": "Утримати керованість CPA.",
        "requests": "",
        "weekly_conversions_reported": 17.0,
        "weekly_cpa_reported": 64.0,
    },
    {
        "week_label": "01.12-07.12",
        "week_start": "2025-12-01",
        "week_end": "2025-12-07",
        "period_label": "in_quarter_week",
        "actions": "Робота з трекінгом.",
        "results": "41 конверсія, CPA 29. Один із найсильніших тижнів кварталу.",
        "plan_to_do": "Масштабувати сильні налаштування без втрати якості.",
        "requests": "",
        "weekly_conversions_reported": 41.0,
        "weekly_cpa_reported": 29.0,
    },
    {
        "week_label": "08.12-14.12",
        "week_start": "2025-12-08",
        "week_end": "2025-12-14",
        "period_label": "in_quarter_week",
        "actions": "Робота з трекінгом.",
        "results": "14 конверсій, CPA 70.",
        "plan_to_do": "Перевірити джерела просідання.",
        "requests": "",
        "weekly_conversions_reported": 14.0,
        "weekly_cpa_reported": 70.0,
    },
    {
        "week_label": "15.12-21.12",
        "week_start": "2025-12-15",
        "week_end": "2025-12-21",
        "period_label": "in_quarter_week",
        "actions": "Робота з трекінгом.",
        "results": "21 конверсія, CPA 77. Попит майже подвоївся, але це не дало пропорційного росту конверсій.",
        "plan_to_do": "Перевірити delayed conversions і прямі замовлення.",
        "requests": "",
        "weekly_conversions_reported": 21.0,
        "weekly_cpa_reported": 77.0,
    },
    {
        "week_label": "22.12-28.12",
        "week_start": "2025-12-22",
        "week_end": "2025-12-28",
        "period_label": "in_quarter_week",
        "actions": "Опрацювання ключових слів.",
        "results": "38 конверсій, CPA 46.",
        "plan_to_do": "Продовжити чистку семантики.",
        "requests": "",
        "weekly_conversions_reported": 38.0,
        "weekly_cpa_reported": 46.0,
    },
    {
        "week_label": "29.12-04.01",
        "week_start": "2025-12-29",
        "week_end": "2026-01-04",
        "period_label": "closing_overlap_week",
        "actions": "Тиждень включено за правилом overlap week.",
        "results": "У квартальному звіті враховано окремо як перехідний тиждень між Q4 2025 і Q1 2026.",
        "plan_to_do": "Тримати окреме маркування в аналітиці.",
        "requests": "",
        "weekly_conversions_reported": "",
        "weekly_cpa_reported": "",
    },
]


tasks = [
    ("2025-10-12", "Тижневий моніторинг", "Done", "2025-10-12", "2025-10-12", 2.0, 2.0, "Monitoring"),
    ("2025-10-19", "Тижневий моніторинг", "Done", "2025-10-19", "2025-10-19", 2.0, 2.0, "Monitoring"),
    ("2025-10-26", "Тижневий моніторинг", "Done", "2025-10-26", "2025-10-26", 2.0, 2.0, "Monitoring"),
    ("2025-11-02", "Тижневий моніторинг", "Done", "2025-11-02", "2025-11-02", 2.0, 2.0, "Monitoring"),
    ("2025-11-09", "Тижневий моніторинг", "Done", "2025-11-09", "2025-11-09", 2.0, 2.0, "Monitoring"),
    ("2025-11-16", "Тижневий моніторинг", "Done", "2025-11-16", "2025-11-16", 2.0, 2.0, "Monitoring"),
    ("2025-11-23", "Тижневий моніторинг", "Done", "2025-11-23", "2026-01-01", 2.0, 2.0, "Monitoring"),
    ("2025-11-30", "Тижневий моніторинг", "Done", "2025-11-30", "2026-01-01", 2.0, 2.0, "Monitoring"),
    ("2025-12-05", "Перенастроить отслеживание юзеров на сайте через Clarity", "Done", "2025-12-05", "2026-01-01", 0.25, 0.25, "Tracking"),
    ("2025-12-07", "Тижневий моніторинг", "Done", "2025-12-07", "2026-01-01", 2.0, 2.0, "Monitoring"),
    ("2025-12-14", "Тижневий моніторинг", "Done", "2025-12-14", "2026-01-01", 2.0, 2.0, "Monitoring"),
    ("2025-12-21", "Тижневий моніторинг", "Done", "2025-12-21", "2025-12-22", 2.0, 2.0, "Monitoring"),
    ("2025-12-28", "Тижневий моніторинг", "Done", "2025-12-28", "2026-01-01", 2.0, 2.0, "Monitoring"),
]


task_rows = [
    {
        "project": "draljundi.com - Afra Fouad",
        "task_name": name,
        "status": status,
        "due_date": due_date,
        "last_task_completion": completion,
        "est_hours": est_hours,
        "real_hours": real_hours,
        "category": category,
        "quarter_attribution": "Q4 2025",
        "attribution_rule": "due_date",
    }
    for _, name, status, due_date, completion, est_hours, real_hours, category in tasks
]


weekly_rows = parse_weekly_rows()
monthly_rows = parse_monthly_rows()
campaign_rows, campaign_totals = parse_single_table(
    BASE / "req81_q4_response.txt",
    ["campaign_status", "campaign_name"],
    ["impressions", "clicks", "ctr", "avg_cpc", "cost", "conversions", "cpa", "conv_rate"],
)
device_rows, device_totals = parse_single_table(
    BASE / "req82_q4_response.txt",
    ["device"],
    ["impressions", "clicks", "ctr", "avg_cpc", "cost", "conversions", "cpa", "conv_rate"],
)
geo_rows, geo_totals = parse_single_table(
    BASE / "req80_q4_response.txt",
    ["region"],
    ["impressions", "clicks", "ctr", "avg_cpc", "cost", "conversions", "cpa", "conv_rate"],
)


total_cost = geo_totals["cost"]
total_conversions = geo_totals["conversions"]
total_impressions = geo_totals["impressions"]
total_clicks = geo_totals["clicks"]
quarter_ctr = geo_totals["ctr"]
quarter_avg_cpc = geo_totals["avg_cpc"]
quarter_cpa = geo_totals["cpa"]
quarter_conv_rate = geo_totals["conv_rate"]

oct_row, nov_row, dec_row = monthly_rows

total_real_hours = sum(r["real_hours"] for r in task_rows)
monitoring_hours = sum(r["real_hours"] for r in task_rows if r["category"] == "Monitoring")
tracking_hours = sum(r["real_hours"] for r in task_rows if r["category"] == "Tracking")

best_week = max(
    [r for r in weekly_reports if isinstance(r["weekly_conversions_reported"], (int, float))],
    key=lambda r: float(r["weekly_conversions_reported"]),
)
worst_cpa_week = max(
    [r for r in weekly_reports if isinstance(r["weekly_cpa_reported"], (int, float))],
    key=lambda r: float(r["weekly_cpa_reported"]),
)

analysis_rows = [
    ("Section", "Metric", "Value", "Notes"),
    ("Executive summary", "Quarter", "Q4 2025", "01 Oct 2025 - 31 Dec 2025"),
    ("Executive summary", "Methodology", "Built from scratch", "Looker Studio + Reports + Task List. Old local exports ignored."),
    ("Executive summary", "Overlap weeks", "Included and labeled", "Opening overlap week 29.09-05.10 and closing overlap week 29.12-04.01 are tracked separately."),
    ("KPI", "Impressions", total_impressions, "Exact Q4 range from Looker Studio"),
    ("KPI", "Clicks", total_clicks, "Exact Q4 range from Looker Studio"),
    ("KPI", "CTR", round2(quarter_ctr * 100), "%"),
    ("KPI", "Avg CPC", round2(quarter_avg_cpc), ""),
    ("KPI", "Cost", round2(total_cost), ""),
    ("KPI", "Conversions", round2(total_conversions), ""),
    ("KPI", "CPA", round2(quarter_cpa), ""),
    ("KPI", "Conversion rate", round2(quarter_conv_rate * 100), "%"),
    ("Monthly trend", "October cost", round2(oct_row["cost"]), ""),
    ("Monthly trend", "October conversions", round2(oct_row["conversions"]), ""),
    ("Monthly trend", "October CPA", round2(oct_row["cpa"]), ""),
    ("Monthly trend", "November cost", round2(nov_row["cost"]), f"{round2(change_pct(nov_row['cost'], oct_row['cost']) * 100)}% vs Oct"),
    ("Monthly trend", "November conversions", round2(nov_row["conversions"]), f"{round2(change_pct(nov_row['conversions'], oct_row['conversions']) * 100)}% vs Oct"),
    ("Monthly trend", "November CPA", round2(nov_row["cpa"]), f"{round2(change_pct(nov_row['cpa'], oct_row['cpa']) * 100)}% vs Oct"),
    ("Monthly trend", "December cost", round2(dec_row["cost"]), f"{round2(change_pct(dec_row['cost'], nov_row['cost']) * 100)}% vs Nov"),
    ("Monthly trend", "December conversions", round2(dec_row["conversions"]), f"{round2(change_pct(dec_row['conversions'], nov_row['conversions']) * 100)}% vs Nov"),
    ("Monthly trend", "December CPA", round2(dec_row["cpa"]), f"{round2(change_pct(dec_row['cpa'], nov_row['cpa']) * 100)}% vs Nov"),
    ("Campaign", "Top spend driver", campaign_rows[0]["campaign_name"], f"{round2(pct(campaign_rows[0]['cost'], total_cost) * 100)}% of spend"),
    ("Campaign", "Top conversion driver", campaign_rows[0]["campaign_name"], f"{round2(pct(campaign_rows[0]['conversions'], total_conversions) * 100)}% of conversions"),
    ("Campaign", "Main inefficiency", campaign_rows[1]["campaign_name"], f"CPA {round2(campaign_rows[1]['cpa'])} with only {round2(campaign_rows[1]['conversions'])} conversions"),
    ("Device", "Primary device", device_rows[0]["device"], f"{round2(pct(device_rows[0]['cost'], total_cost) * 100)}% of spend, {round2(device_rows[0]['conversions'])} conversions"),
    ("Geo", "Primary geo", geo_rows[0]["region"], "Single visible geo bucket in Q4 export"),
    ("Work completed", "Weekly reports captured", len(weekly_reports), "13 weekly reports plus opening overlap week"),
    ("Work completed", "Completed tasks captured", len(task_rows), "Attributed by due date, not by backfilled completion timestamp"),
    ("Work completed", "Real hours", round2(total_real_hours), "24.0 monitoring hours + 0.25 tracking hours"),
    ("Work completed", "Best reported week", best_week["week_label"], best_week["results"]),
    ("Work completed", "Most difficult reported week", worst_cpa_week["week_label"], worst_cpa_week["results"]),
]


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


for filename, rows in [
    ("q4_2025_raw_looker_weekly.csv", weekly_rows),
    ("q4_2025_raw_looker_monthly.csv", monthly_rows),
    ("q4_2025_raw_campaign.csv", campaign_rows),
    ("q4_2025_raw_device.csv", device_rows),
    ("q4_2025_raw_geo.csv", geo_rows),
    ("q4_2025_raw_weekly_reports.csv", weekly_reports),
    ("q4_2025_raw_tasks.csv", task_rows),
]:
    write_csv(BASE / filename, rows)


wb = Workbook()
default = wb.active
wb.remove(default)


def add_sheet(name: str, rows: list[dict] | list[tuple]) -> None:
    ws = wb.create_sheet(title=name)
    if not rows:
        return
    if isinstance(rows[0], dict):
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    else:
        for row in rows:
            ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 48)


add_sheet("Raw - Looker Weekly", weekly_rows)
add_sheet("Raw - Looker Monthly", monthly_rows)
add_sheet("Raw - Campaign", campaign_rows)
add_sheet("Raw - Device", device_rows)
add_sheet("Raw - Geo", geo_rows)
add_sheet("Raw - Weekly Reports", weekly_reports)
add_sheet("Raw - Tasks", task_rows)
add_sheet("Analysis", analysis_rows)

xlsx_path = BASE / "q4_2025_draljundi_afra_fouad_report.xlsx"
wb.save(xlsx_path)


report_md = f"""# Quarterly Report: draljundi.com / Afra Fouad — Q4 2025

## Executive summary

Звіт побудовано з нуля за період **01.10.2025 - 31.12.2025** на основі трьох джерел:

- Looker Studio
- weekly reports у `Reports`
- operational tasks у `Task List`

Попередні локальні вигрузки та старі аналізи не використовувалися як джерело даних.

Ключовий результат кварталу: при витратах **{round2(total_cost)}** акаунт отримав **{round2(total_conversions)}** конверсій з **CPA {round2(quarter_cpa)}**. Найсильнішим місяцем став грудень: **{round2(dec_row['conversions'])}** конверсій при **CPA {round2(dec_row['cpa'])}**, що суттєво краще за листопад.

## KPI summary

- Impressions: **{total_impressions:,}**
- Clicks: **{total_clicks:,}**
- CTR: **{round2(quarter_ctr * 100)}%**
- Avg CPC: **{round2(quarter_avg_cpc)}**
- Cost: **{round2(total_cost)}**
- Conversions: **{round2(total_conversions)}**
- CPA: **{round2(quarter_cpa)}**
- Conversion rate: **{round2(quarter_conv_rate * 100)}%**

## Performance analysis

### Monthly trend

- **October 2025**: {round2(oct_row['conversions'])} conversions, cost {round2(oct_row['cost'])}, CPA {round2(oct_row['cpa'])}
- **November 2025**: {round2(nov_row['conversions'])} conversions, cost {round2(nov_row['cost'])}, CPA {round2(nov_row['cpa'])}
- **December 2025**: {round2(dec_row['conversions'])} conversions, cost {round2(dec_row['cost'])}, CPA {round2(dec_row['cpa'])}

Інтерпретація:

- Листопад просів проти жовтня за кількістю конверсій і погіршився за CPA.
- Грудень став точкою відновлення: витрати залишилися майже на рівні листопада, але кількість конверсій різко зросла.
- Це свідчить, що зміни в оптимізації, таргетингу та роботі з семантикою дали ефект ближче до кінця кварталу.

### Campaign analysis

- **{campaign_rows[0]['campaign_name']}** був основним драйвером кварталу: {round2(campaign_rows[0]['conversions'])} конверсій, CPA {round2(campaign_rows[0]['cpa'])}, частка витрат {round2(pct(campaign_rows[0]['cost'], total_cost) * 100)}%.
- **{campaign_rows[1]['campaign_name']}** споживав помітний бюджет, але дав лише {round2(campaign_rows[1]['conversions'])} конверсії при CPA {round2(campaign_rows[1]['cpa'])}; це головна зона для перегляду в наступному кварталі.
- **{campaign_rows[2]['campaign_name']}** мав низький обсяг, але кращий CPA; це скоріше точковий допоміжний канал, ніж масштабований драйвер.

### Device analysis

- Майже вся ефективність кварталу припала на **mobile devices with full browsers**: {round2(device_rows[0]['conversions'])} конверсій і {round2(pct(device_rows[0]['cost'], total_cost) * 100)}% витрат.
- Desktop і tablet трафік були незначними за обсягом і не дали конверсій у Q4.
- Практичний висновок: mobile-first оптимізація має лишатися базовим сценарієм для подальшої роботи.

### Geo analysis

- У Q4-експорті вся видима активність агрегується в **{geo_rows[0]['region']}**.
- Це спрощує контроль якості трафіку, але звужує простір для geo-диференціації. Якщо в акаунті з’явиться деталізація за містами або районами, її варто винести в окремий зріз у наступному кварталі.

## Work completed

### What the team actually did

За weekly reports у кварталі зафіксовано такі типи робіт:

- управління бюджетом і темпом витрат
- зміни bidding strategy
- опрацювання ключових слів і negative keywords
- geo exclusions
- розширення language targeting
- оновлення PMax asset groups
- регулярна перевірка трекінгу
- моніторинг аномалій і перевірка на click fraud

### Task control layer

Task List підтверджує регулярний операційний ритм:

- завершено **{len(task_rows)}** задач, атрибутованих до Q4 2025
- сумарно зафіксовано **{round2(total_real_hours)}** real hours
- із них **{round2(monitoring_hours)}** годин припадає на weekly monitoring
- окремо зафіксовано технічну задачу з **Clarity tracking**

Важлива методологічна примітка: частина задач із due date у листопаді та грудні має `Last Task Completion` у січні 2026. Для квартального звіту задачі атрибутовано до Q4 за **due date**, а не за пізнішим технічним timestamp завершення.

### Weekly execution highlights

- Найкращі тижні кварталу за weekly reports: `06.10-12.10`, `27.10-02.11`, `01.12-07.12`, `22.12-28.12`
- Найслабші тижні за reported CPA: `13.10-19.10`, `03.11-09.11`, `15.12-21.12`
- Перехідний тиждень `29.12.2025 - 04.01.2026` включено в окремий блок як **overlap week**

## Wins / issues / anomalies

### Wins

- У грудні вдалося вийти на найкращий місячний результат кварталу без суттєвого росту витрат.
- Основний обсяг конверсій стабільно давав PMax-контур, що спростило фокус оптимізації.
- Регулярний weekly monitoring тримав акаунт у робочому ритмі протягом усього кварталу.

### Issues

- Search/English-напрямок спалював бюджет суттєво гірше за основну PMax-кампанію.
- Частина weekly reports показувала різкі стрибки CPA, що вказує на чутливість до сезонного попиту та/або якості трафіку.
- Operational logging у Task List частково закривався із запізненням, тому для квартальної атрибуції потрібне правило `due date first`.

### Anomalies

- Opening week `29.09-05.10` містить тільки частину Q4 і маркується як opening overlap.
- Closing week `29.12-04.01` частково належить уже до Q1 2026 і теж маркується окремо.

## Recommendations for next quarter

1. Зберегти PMax як головний масштабуючий контур, але винести окремий контроль якості lead/conversion mix.
2. Переглянути доцільність бюджету на `ADa - Eng`: або жорстко перезібрати, або обмежити, якщо роль кампанії не змінилась.
3. Продовжити mobile-first оптимізацію посадкових і креативів, бо саме mobile формує майже весь результат.
4. Формалізувати logging у weekly tasks: completion timestamp має відповідати фактичному тижню виконання, щоб не виникало backfill-ефекту.
5. Для наступного кварталу додати в Looker окремий стабільний city/placement breakdown, якщо джерело дозволяє детальніший розріз.

## Sources and deliverables

- Project page: https://www.notion.so/c644e93facb64d27b61594d4dd4ef9ef
- Looker Studio: https://lookerstudio.google.com/reporting/b383c1ff-1897-4420-ad45-215cab8482fa/page/p_2rqnedbhbd
- Reports data source: collection://01f2d087-824e-40fc-9ed8-d4c405de6f66
- Task List data source: collection://82ace225-03de-4c1c-b67c-c50b310a8527
- Google Sheets workbook: [to be inserted after upload]

## Methodology note

Цей звіт побудований **з нуля**. Для фінальних KPI використано точний Q4 date range у Looker Studio (`2025-10-01` .. `2025-12-31`). Weekly reports та task data використані як narrative/control layer. Перехідні тижні позначені окремо, щоб не змішувати квартальний зріз із календарною тижневою логікою.
"""

(BASE / "q4_2025_report_draft.md").write_text(report_md, encoding="utf-8")
print(f"Workbook: {xlsx_path}")
print(f"Draft report: {BASE / 'q4_2025_report_draft.md'}")
