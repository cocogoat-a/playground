#!/usr/bin/env python3
"""
Build one cleaned quarterly Google Ads workbook from native section exports.

Expected input filenames inside --input-dir:
  current_quarter_<report_id>.csv
  last_quarter_<report_id>.csv

Supported report_id values:
  ad_group_performance
  devices
  locations
  time_of_day
  age
  gender
  household_income
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORTS = [
    ("ad_group_performance", "Ad Group Performance"),
    ("devices", "Devices"),
    ("locations", "Locations"),
    ("time_of_day", "Time of Day"),
    ("age", "Age"),
    ("gender", "Gender"),
    ("household_income", "Household Income"),
]

TITLE_ROWS = {
    "Campaign report",
    "Ad group report",
    "Device report",
    "Matched locations report",
    "Ad schedule day and hour report",
}

NUMERIC_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^-?\d+(?:\.\d+)?$")
PERCENT_RE = re.compile(r"^-?\d+(?:\.\d+)?%$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--account-name", default="")
    parser.add_argument("--account-cid", default="")
    parser.add_argument("--current-quarter-label", default="Current Quarter")
    parser.add_argument("--last-quarter-label", default="Last Quarter")
    parser.add_argument("--current-quarter-range", default="")
    parser.add_argument("--last-quarter-range", default="")
    return parser.parse_args()


def is_blank_row(row: list[str]) -> bool:
    return not row or all(not str(cell).strip() for cell in row)


def is_total_row(row: list[str]) -> bool:
    for cell in row:
        text = str(cell).strip()
        if text.startswith("Total:"):
            return True
    return False


def normalize_row(row: list[str]) -> list[str]:
    normalized = [str(cell).strip() for cell in row]
    while normalized and normalized[-1] == "":
        normalized.pop()
    return normalized


def detect_header_row(rows: list[list[str]]) -> int:
    for idx, row in enumerate(rows):
        if is_blank_row(row):
            continue
        if row[0].strip() in TITLE_ROWS:
            continue
        if idx + 1 < len(rows) and re.search(r"\d", ",".join(rows[idx + 1])):
            return idx
    raise ValueError("Could not detect header row")


def convert_cell(value: str):
    text = value.strip()
    if text in {"", " --"}:
        return None
    if text in {"—", "--"}:
        return text
    if PERCENT_RE.match(text):
        return float(text[:-1]) / 100.0
    if NUMERIC_RE.match(text):
        return float(text.replace(",", ""))
    return text


def find_header_index(header: list[str], names: set[str]) -> int | None:
    lowered = [str(cell).strip().lower() for cell in header]
    for idx, cell in enumerate(lowered):
        if cell in names:
            return idx
    return None


def is_zero_useful_stats_row(header: list[str], row: list[object]) -> bool:
    impressions_idx = find_header_index(header, {"impr.", "impressions"})
    cost_idx = find_header_index(header, {"cost"})
    if impressions_idx is None or cost_idx is None:
        return False

    impressions = row[impressions_idx]
    cost = row[cost_idx]
    if not isinstance(impressions, (int, float)) or not isinstance(cost, (int, float)):
        return False
    return impressions == 0 and cost == 0


def clean_csv(path: Path, quarter_label: str) -> tuple[list[str], list[list[object]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        raw_rows = [normalize_row(row) for row in csv.reader(handle)]

    raw_rows = [row for row in raw_rows if not is_blank_row(row)]
    header_idx = detect_header_row(raw_rows)
    header = raw_rows[header_idx]
    cleaned_rows: list[list[object]] = []

    for row in raw_rows[header_idx + 1 :]:
        if is_blank_row(row) or is_total_row(row):
            continue
        row = row[: len(header)] + [""] * max(0, len(header) - len(row))
        converted_row = [convert_cell(cell) for cell in row]
        if is_zero_useful_stats_row(header, converted_row):
            continue
        cleaned_rows.append([quarter_label] + converted_row)

    return ["Quarter"] + header, cleaned_rows


def autosize_columns(sheet) -> None:
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def apply_number_formats(sheet) -> None:
    headers = {cell.column: str(cell.value) for cell in sheet[1]}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if not isinstance(cell.value, float):
                continue
            header = headers.get(cell.column, "").lower()
            if "rate" in header:
                cell.number_format = "0.00%"
            elif any(token in header for token in ["cost", "budget", "cpc"]):
                cell.number_format = "#,##0.00"
            else:
                cell.number_format = "#,##0.00"


def write_summary(
    sheet,
    account_name: str,
    account_cid: str,
    current_quarter_range: str,
    last_quarter_range: str,
    available_tabs: list[str],
    missing_reports: list[str],
) -> None:
    rows = [
        ("Account Name", account_name or None),
        ("Account CID", account_cid or None),
        ("Run Date", str(date.today())),
        ("Current Quarter", current_quarter_range or None),
        ("Last Quarter", last_quarter_range or None),
        ("Included Tabs", ", ".join(available_tabs) or None),
        ("Missing Reports", ", ".join(missing_reports) or "None"),
    ]
    sheet["A1"] = "Google Ads Quarterly Export Summary"
    sheet["A1"].font = Font(bold=True, size=14)
    for idx, (label, value) in enumerate(rows, start=3):
        sheet[f"A{idx}"] = label
        sheet[f"A{idx}"].font = Font(bold=True)
        sheet[f"B{idx}"] = value
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 80


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_path = Path(args.output_xlsx)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("Summary")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    available_tabs: list[str] = []
    missing_reports: list[str] = []

    for report_id, tab_name in REPORTS:
        combined_headers = None
        combined_rows: list[list[object]] = []

        for prefix, quarter_label in [
            ("current_quarter", args.current_quarter_label),
            ("last_quarter", args.last_quarter_label),
        ]:
            source = input_dir / f"{prefix}_{report_id}.csv"
            if not source.exists():
                continue
            headers, rows = clean_csv(source, quarter_label)
            combined_headers = headers
            combined_rows.extend(rows)

        if not combined_headers:
            missing_reports.append(tab_name)
            continue

        sheet = workbook.create_sheet(tab_name)
        available_tabs.append(tab_name)
        for col_idx, value in enumerate(combined_headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(combined_rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        apply_number_formats(sheet)
        autosize_columns(sheet)

    write_summary(
        summary,
        args.account_name,
        args.account_cid,
        args.current_quarter_range,
        args.last_quarter_range,
        available_tabs,
        missing_reports,
    )

    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
