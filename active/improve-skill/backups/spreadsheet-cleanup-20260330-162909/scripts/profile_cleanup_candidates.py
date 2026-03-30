#!/usr/bin/env python3
"""
Profile a local spreadsheet and report candidate low-signal rows and columns.

Supported formats:
- .csv
- .tsv
- .xlsx (requires openpyxl)
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


PLACEHOLDERS = {
    "",
    "--",
    "[]",
    "none",
    "null",
    "n/a",
    "na",
    "nil",
    "(not set)",
    "—",
}

KEEP_IF_CONSTANT_TERMS = {
    "quarter",
    "date",
    "day",
    "hour",
    "week",
    "month",
    "year",
    "campaign",
    "ad set",
    "adset",
    "ad group",
    "adgroup",
    "ad",
    "device",
    "location",
    "country",
    "region",
    "state",
    "city",
    "gender",
    "age",
    "audience",
    "segment",
    "placement",
}

LOW_SIGNAL_HEADER_PATTERNS = (
    "status",
    "status reason",
    "currency code",
    "target cpa",
    "target roas",
    "default max cpc",
    "bid adj",
    "level",
    "brand inclusions",
)

LOW_SIGNAL_CONSTANT_VALUES = {
    "enabled",
    "eligible",
    "not eligible",
    "campaign",
    "ad group",
    "none",
    "unknown",
}

DIRECT_COST_HEADERS = {
    "cost",
    "spend",
    "amount spent",
    "ad spend",
    "costs",
}

DIRECT_CONVERSION_HEADERS = {
    "conversions",
    "conversion",
    "conv",
    "conv platform comparable",
    "results",
    "purchases",
    "purchase",
}

NON_DIRECT_TOKENS = {
    "avg",
    "average",
    "rate",
    "value",
    "per",
    "cpa",
    "roas",
    "ctr",
    "cpc",
    "cpm",
}


@dataclass
class ColumnProfile:
    index: int
    header: str
    normalized_header: str
    distinct_non_placeholder: list[str]
    distinct_count: int
    placeholder_count: int
    total_count: int
    candidate_reason: str | None = None
    duplicate_of: str | None = None


@dataclass
class SheetProfile:
    name: str
    header_row_number: int
    source_rows: int
    data_rows: int
    cost_column: str | None
    conversions_column: str | None
    removable_rows: list[int]
    candidate_columns: list[ColumnProfile]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_header(value: object) -> str:
    text = normalize_text(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_placeholder(value: object) -> bool:
    return normalize_text(value).lower() in PLACEHOLDERS


def parse_number(value: object) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    lower = text.lower()
    if lower in PLACEHOLDERS:
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    text = text.replace("%", "")
    text = re.sub(r"[^\d,.\-]", "", text)
    if not text:
        return None

    if "," in text and "." in text:
        text = text.replace(",", "")
    elif text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")

    try:
        number = float(text)
    except ValueError:
        return None

    if negative:
        number = -number
    if math.isfinite(number):
        return number
    return None


def column_letter(index: int) -> str:
    index += 1
    chars: list[str] = []
    while index:
        index, remainder = divmod(index - 1, 26)
        chars.append(chr(65 + remainder))
    return "".join(reversed(chars))


def first_non_empty_row(rows: list[list[object]]) -> int:
    for idx, row in enumerate(rows):
        if any(normalize_text(cell) for cell in row):
            return idx
    return 0


def pad_rows(rows: list[list[object]]) -> list[list[object]]:
    width = max((len(row) for row in rows), default=0)
    return [row + [""] * (width - len(row)) for row in rows]


def read_delimited(path: Path, delimiter: str) -> list[tuple[str, list[list[object]]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = [list(row) for row in reader]
    return [(path.stem, pad_rows(rows))]


def read_xlsx(path: Path, requested_sheet: str | None) -> list[tuple[str, list[list[object]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for .xlsx files. Install it with `uv pip install openpyxl`."
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_names = [requested_sheet] if requested_sheet else workbook.sheetnames
    results: list[tuple[str, list[list[object]]]] = []
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"Sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        results.append((sheet_name, pad_rows(rows)))
    return results


def read_source(path: Path, requested_sheet: str | None) -> list[tuple[str, list[list[object]]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_delimited(path, ",")
    if suffix == ".tsv":
        return read_delimited(path, "\t")
    if suffix == ".xlsx":
        return read_xlsx(path, requested_sheet)
    raise SystemExit(f"Unsupported file type: {path.suffix}")


def is_direct_cost_header(normalized: str) -> bool:
    if normalized in DIRECT_COST_HEADERS:
        return True
    tokens = set(normalized.split())
    if "cost" in tokens or "spend" in tokens:
        return not (tokens & NON_DIRECT_TOKENS or "conv" in tokens or "conversion" in tokens)
    return False


def is_direct_conversions_header(normalized: str) -> bool:
    if normalized in DIRECT_CONVERSION_HEADERS:
        return True
    tokens = set(normalized.split())
    has_conversion_word = bool(tokens & {"conv", "conversion", "conversions", "result", "results", "purchase", "purchases"})
    if not has_conversion_word:
        return False
    if tokens & {"cost", "spend", "value"}:
        return False
    return not bool(tokens & NON_DIRECT_TOKENS)


def detect_metric_columns(headers: list[str]) -> tuple[int | None, int | None]:
    cost_idx = None
    conv_idx = None
    for idx, header in enumerate(headers):
        normalized = normalize_header(header)
        if cost_idx is None and is_direct_cost_header(normalized):
            cost_idx = idx
            continue
        if conv_idx is None and is_direct_conversions_header(normalized):
            conv_idx = idx
    return cost_idx, conv_idx


def should_keep_constant_column(normalized_header: str) -> bool:
    return any(term in normalized_header for term in KEEP_IF_CONSTANT_TERMS)


def is_low_signal_constant_column(normalized_header: str, value: str) -> bool:
    if should_keep_constant_column(normalized_header):
        return False
    if any(pattern in normalized_header for pattern in LOW_SIGNAL_HEADER_PATTERNS):
        return True
    return value.lower() in LOW_SIGNAL_CONSTANT_VALUES


def signature_for_values(values: Iterable[object]) -> tuple[str, ...]:
    return tuple(normalize_text(value).lower() for value in values)


def profile_sheet(sheet_name: str, rows: list[list[object]]) -> SheetProfile:
    if not rows:
        return SheetProfile(sheet_name, 1, 0, 0, None, None, [], [])

    header_idx = first_non_empty_row(rows)
    padded = pad_rows(rows)
    headers = [normalize_text(value) for value in padded[header_idx]]
    data_rows = padded[header_idx + 1 :]
    cost_idx, conv_idx = detect_metric_columns(headers)

    removable_rows: list[int] = []
    kept_rows: list[list[object]] = []
    for offset, row in enumerate(data_rows, start=header_idx + 2):
        if cost_idx is not None and conv_idx is not None:
            cost = parse_number(row[cost_idx])
            conversions = parse_number(row[conv_idx])
            if cost is not None and conversions is not None and cost == 0 and conversions == 0:
                removable_rows.append(offset)
                continue
        kept_rows.append(row)

    headers_seen: dict[str, str] = {}
    signatures_seen: dict[tuple[str, ...], str] = {}
    candidate_columns: list[ColumnProfile] = []

    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        values = [row[index] if index < len(row) else "" for row in kept_rows]
        normalized_values = [normalize_text(value) for value in values]
        non_placeholder = [value for value in normalized_values if not is_placeholder(value)]
        distinct_non_placeholder = sorted(set(non_placeholder))
        placeholder_count = sum(1 for value in values if is_placeholder(value))
        signature = signature_for_values(values)

        profile = ColumnProfile(
            index=index,
            header=header or f"Column {column_letter(index)}",
            normalized_header=normalized,
            distinct_non_placeholder=distinct_non_placeholder,
            distinct_count=len(distinct_non_placeholder),
            placeholder_count=placeholder_count,
            total_count=len(values),
        )

        if len(non_placeholder) == 0:
            profile.candidate_reason = "blank-or-placeholder-only"
        elif normalized in headers_seen:
            profile.candidate_reason = "duplicate-header"
            profile.duplicate_of = headers_seen[normalized]
        elif signature in signatures_seen and (normalized == normalize_header(signatures_seen[signature]) or len(distinct_non_placeholder) <= 1):
            profile.candidate_reason = "duplicate-data"
            profile.duplicate_of = signatures_seen[signature]
        elif len(distinct_non_placeholder) == 1 and is_low_signal_constant_column(normalized, distinct_non_placeholder[0]):
            profile.candidate_reason = "constant-low-signal"

        if normalized and normalized not in headers_seen:
            headers_seen[normalized] = profile.header
        if signature not in signatures_seen:
            signatures_seen[signature] = profile.header

        if profile.candidate_reason:
            candidate_columns.append(profile)

    cost_name = headers[cost_idx] if cost_idx is not None else None
    conv_name = headers[conv_idx] if conv_idx is not None else None

    return SheetProfile(
        name=sheet_name,
        header_row_number=header_idx + 1,
        source_rows=len(rows),
        data_rows=len(data_rows),
        cost_column=cost_name,
        conversions_column=conv_name,
        removable_rows=removable_rows,
        candidate_columns=candidate_columns,
    )


def render_text_report(path: Path, profiles: list[SheetProfile]) -> str:
    lines = [f"Source: {path}"]
    for profile in profiles:
        lines.append("")
        lines.append(f"Sheet: {profile.name}")
        lines.append(f"Header row: {profile.header_row_number}")
        lines.append(f"Data rows: {profile.data_rows}")
        lines.append(f"Detected cost column: {profile.cost_column or 'none'}")
        lines.append(f"Detected conversions column: {profile.conversions_column or 'none'}")
        if profile.removable_rows:
            preview = ", ".join(str(row) for row in profile.removable_rows[:20])
            suffix = "" if len(profile.removable_rows) <= 20 else ", ..."
            lines.append(
                f"Candidate rows to remove ({len(profile.removable_rows)}): {preview}{suffix}"
            )
        else:
            lines.append("Candidate rows to remove: none")

        if profile.candidate_columns:
            lines.append("Candidate columns to remove:")
            for column in profile.candidate_columns:
                value_preview = ", ".join(column.distinct_non_placeholder[:3]) or "no non-placeholder values"
                duplicate = f" -> {column.duplicate_of}" if column.duplicate_of else ""
                lines.append(
                    f"  - {column_letter(column.index)} `{column.header}`: {column.candidate_reason}{duplicate}; values: {value_preview}"
                )
        else:
            lines.append("Candidate columns to remove: none")
    return "\n".join(lines)


def render_json_report(path: Path, profiles: list[SheetProfile]) -> str:
    payload = {
        "source": str(path),
        "sheets": [
            {
                "name": profile.name,
                "header_row": profile.header_row_number,
                "data_rows": profile.data_rows,
                "cost_column": profile.cost_column,
                "conversions_column": profile.conversions_column,
                "candidate_rows_to_remove": profile.removable_rows,
                "candidate_columns_to_remove": [
                    {
                        "index": column.index,
                        "letter": column_letter(column.index),
                        "header": column.header,
                        "reason": column.candidate_reason,
                        "duplicate_of": column.duplicate_of,
                        "distinct_non_placeholder": column.distinct_non_placeholder,
                    }
                    for column in profile.candidate_columns
                ],
            }
            for profile in profiles
        ],
    }
    return json.dumps(payload, indent=2)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Profile spreadsheets and report low-signal rows and columns."
    )
    parser.add_argument("path", help="Path to a local .csv, .tsv, or .xlsx file")
    parser.add_argument("--sheet", help="Analyze a single worksheet in an .xlsx file")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Emit a JSON report instead of plain text",
    )
    args = parser.parse_args()

    path = Path(args.path).resolve()
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    raw_sheets = read_source(path, args.sheet)
    profiles = [profile_sheet(name, rows) for name, rows in raw_sheets]

    output = render_json_report(path, profiles) if args.json else render_text_report(path, profiles)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
